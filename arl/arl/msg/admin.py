from uuid import uuid4

import boto3
from django.conf import settings
from django.contrib import admin
from django.http import HttpResponse
from django.utils.html import format_html
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from arl.bucket.helpers import upload_to_linode_object_storage  # adjust if needed

from .models import ComplianceFile, EmailEvent


@admin.register(ComplianceFile)
class ComplianceFileAdmin(admin.ModelAdmin):
    list_display = ("title", "uploaded_at", "is_active", "presigned_link_display")
    list_filter = ("is_active",)
    search_fields = ("title",)
    readonly_fields = ("s3_key", "presigned_url", "uploaded_at")

    def presigned_link_display(self, obj):
        if obj.presigned_url:
            return format_html(
                '<a href="{}" target="_blank">View File</a>', obj.presigned_url
            )
        return "-"

    presigned_link_display.short_description = "Download Link"

    def save_model(self, request, obj, form, change):
        if obj.is_active:
            ComplianceFile.objects.exclude(pk=obj.pk).update(is_active=False)

        super().save_model(request, obj, form, change)

        if obj.file and not obj.s3_key:
            key = f"compliance/{uuid4()}_{obj.file.name}"
            upload_to_linode_object_storage(obj.file.file, key)
            obj.s3_key = key

            # âœ… Generate new presigned URL
            conn = boto3.client(
                "s3",
                endpoint_url=settings.LINODE_ENDPOINT,
                aws_access_key_id=settings.LINODE_ACCESS_KEY,
                aws_secret_access_key=settings.LINODE_SECRET_KEY,
                region_name="us-east-1",
            )
            obj.presigned_url = conn.generate_presigned_url(
                ClientMethod="get_object",
                Params={"Bucket": settings.LINODE_BUCKET_NAME, "Key": key},
                ExpiresIn=60 * 60 * 24 * 7,
            )

            obj.save(update_fields=["s3_key", "presigned_url"])


@admin.register(EmailEvent)
class EmailEventAdmin(admin.ModelAdmin):
    list_display = (
        "timestamp",
        "email",
        "event",
        "subject",
        "sg_template_name",
        "employer",
    )
    list_filter = ("event", "employer", "sg_template_name", "timestamp")
    search_fields = (
        "email",
        "subject",
        "sg_template_name",
        "sg_message_id",
        "sg_event_id",
    )
    ordering = ("-timestamp",)
    actions = ["export_selected_to_xlsx"]

    def export_selected_to_xlsx(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "Email Events"

        headers = [
            "Timestamp",
            "Email",
            "Event",
            "Subject",
            "Template Name",
            "Template ID",
            "Message ID",
            "Event ID",
            "Employer",
            "Username",
            "IP",
            "URL",
            "User Agent",
        ]
        ws.append(headers)

        for obj in queryset.iterator():
            ws.append(
                [
                    obj.timestamp.isoformat() if obj.timestamp else "",
                    obj.email,
                    obj.event,
                    obj.subject or "",
                    obj.sg_template_name or "",
                    obj.sg_template_id or "",
                    obj.sg_message_id or "",
                    obj.sg_event_id or "",
                    getattr(obj.employer, "name", "") if obj.employer else "",
                    obj.username or "",
                    obj.ip or "",
                    obj.url or "",
                    (obj.useragent or ""),
                ]
            )

        # autosize-ish
        for col_idx, _ in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 22

        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = "attachment; filename=email_events.xlsx"
        wb.save(resp)
        return resp

    export_selected_to_xlsx.short_description = "Export selected EmailEvents to XLSX"
